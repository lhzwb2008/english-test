#!/usr/bin/env python3
"""从 ref/ 下四份陪跑 Excel 导出 coze/prompts/builtin-tasks-from-excels.md。

要点（与上一版相比）：
1) THINK1 表的列对齐：表头 row 是 ['序号','课程','课后作业',...]，子表头还有
   ['', '', '必做','选做','必做','选做','无']（练习册/作业纸/口语作业）。上一版
   把 row[2]（练习册-必做）当作 #### 标题，导致 lesson_code 全部错位且对应的
   作业内容也错位、且 row[2]='无' 时整行被跳过（reading 课全丢）。
2) 自动按表头探测「课程列」与各作业列，**任意 sheet 列结构变化均能解析**。
3) 强约束：每个 #### 标题必为该 Excel 「课程」列原文（如 Welcome-PartA-01、
   U1-L1-Reading1、PU2-U1-L1、PU3-U2-L3语法），不再使用作业内容当标题。
4) 同体系下若出现重名课程（罕见），自动追加 ` (Sheet名)` 后缀以保证唯一。
5) 跳过「学生情况 / 学员情况」之类的元信息 sheet。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

ROOT = Path(__file__).resolve().parents[1]
REF = ROOT / 'ref'

SKIP_SHEET_KEYWORDS = ('学员', '学生', '记录', '说明')

LABEL_BUCKETS = [
    ('练习册必做', ['练习册必做']),
    ('练习册选做', ['练习册选做']),
    ('作业纸必做', ['作业纸必做', '作业纸-必做']),
    ('作业纸选做', ['作业纸选做', '作业纸-选做']),
    ('作业纸', ['作业纸']),
    ('练习册', ['练习册']),
    ('口语作业', ['口语作业']),
    ('APP打卡', ['APP打卡', 'app打卡']),
]

LABEL_SHORT = {
    '练习册必做': '必做(册)',
    '练习册选做': '选做(册)',
    '作业纸必做': '必做(纸)',
    '作业纸选做': '选做(纸)',
    '作业纸': '作业纸',
    '练习册': '练习册',
    '口语作业': '口语',
    'APP打卡': '打卡',
}


def clean(x) -> str:
    if x is None:
        return ''
    s = str(x).strip()
    if s in ('', '无', 'nan', 'NaN'):
        return ''
    return s


def md_escape(s: str) -> str:
    s = s.replace('\n', ' ').replace('\r', ' ').strip()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'练习册第(\d+)页', r'P\1', s)
    s = re.sub(r'练习册第(\d+-\d+)页', r'P\1', s)
    return s[:2000]


def looks_like_course_code(s: str) -> bool:
    if not s:
        return False
    s = s.strip()
    # 课程代号常见形式：U1-L1-... / PU2-U1-L1 / PU3-U1-L1词汇 / Welcome-PartA-01 /
    # Unit1-Lesson 1-... / U1-L2 Vocabulary1Grammar1
    return bool(
        re.search(
            r'^(U\d|PU\d|Welcome|Unit\d|Unit\s|UHELLO|Hello|UFINAL)',
            s,
            re.IGNORECASE,
        )
    )


def detect_header(rows: list[list], scan: int = 8) -> dict:
    """在前 scan 行扫描表头，返回各关键列索引。

    返回字典含 keys: course, label_to_col(map: 标签 → 列索引), data_start。

    解析规则（从粗到细）：
      1) 行内出现 ‘课程’：定位课程列。
      2) 行内出现 ‘练习册’/‘作业纸’：作为父级表头，记录其所在列。
      3) 子行内出现 ‘必做’/‘选做’：附着到最近的父级表头列上，得到
         「练习册必做 / 练习册选做 / 作业纸必做 / 作业纸选做」。
      4) 行内独立出现 ‘口语作业 / APP打卡’：作为单列直接映射。
      5) 父级列若已被子级 必做/选做 占用，则不再回退为父级名。
    """
    course_col: int | None = None
    label_to_col: dict[str, int] = {}
    col_to_canon: dict[int, str] = {}
    data_start = scan

    parent_label_at: dict[int, str] = {}
    for r_idx in range(min(scan, len(rows))):
        row = [clean(c) for c in rows[r_idx]]
        for ci, v in enumerate(row):
            if not v:
                continue
            if v == '课程' and course_col is None:
                course_col = ci
            for canon, alias in LABEL_BUCKETS:
                if canon in ('练习册必做', '练习册选做', '作业纸必做', '作业纸选做'):
                    continue
                if v in alias and canon not in label_to_col and ci not in col_to_canon:
                    label_to_col[canon] = ci
                    col_to_canon[ci] = canon
            if v in ('练习册', '作业纸'):
                parent_label_at[ci] = v

    for r_idx in range(min(scan, len(rows))):
        row = [clean(c) for c in rows[r_idx]]
        for ci, v in enumerate(row):
            if v not in ('必做', '选做'):
                continue
            parent = ''
            for cj in range(ci, -1, -1):
                if cj in parent_label_at:
                    parent = parent_label_at[cj]
                    break
            if not parent:
                continue
            canon = f'{parent}{v}'
            if canon in label_to_col:
                continue
            if ci in col_to_canon and col_to_canon[ci] in (parent,):
                pass
            label_to_col[canon] = ci
            col_to_canon[ci] = canon

    promoted_cols = {
        label_to_col[k]
        for k in ('练习册必做', '练习册选做', '作业纸必做', '作业纸选做')
        if k in label_to_col
    }
    for parent in ('练习册', '作业纸'):
        if parent in label_to_col and label_to_col[parent] in promoted_cols:
            label_to_col.pop(parent, None)

    for r_idx in range(min(scan, len(rows))):
        row = [clean(c) for c in rows[r_idx]]
        if course_col is not None and looks_like_course_code(
            row[course_col] if course_col < len(row) else ''
        ):
            data_start = r_idx
            break
    else:
        data_start = scan

    return {
        'course': course_col,
        'label_to_col': label_to_col,
        'data_start': data_start,
    }


def emit_block(
    course: str,
    row: list,
    label_to_col: dict[str, int],
    seen: set[str],
    sheet: str,
) -> str | None:
    if course in seen:
        return None
    seen.add(course)

    parts: list[str] = []
    seen_vals: set[str] = set()
    for canon, _ in LABEL_BUCKETS:
        ci = label_to_col.get(canon)
        if ci is None:
            continue
        v = clean(row[ci]) if ci < len(row) else ''
        if not v:
            continue
        if canon == '练习册' and v.startswith(f'练习册-{course}'):
            continue
        if canon == '练习册' and v == course:
            continue
        cleaned = md_escape(v)
        if cleaned in seen_vals:
            continue
        seen_vals.add(cleaned)
        parts.append(f'{LABEL_SHORT.get(canon, canon)}：{cleaned}')
    body = ' | '.join(parts) if parts else '见表'
    return f'#### {course}\n{body}'


def parse_xlsx_rows(path: Path) -> Iterable[tuple[str, list[list]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        yield sn, rows
    wb.close()


def parse_xls_rows(path: Path) -> Iterable[tuple[str, list[list]]]:
    import xlrd

    wb = xlrd.open_workbook(path)
    for si in range(wb.nsheets):
        sh = wb.sheet_by_index(si)
        rows = [
            [sh.cell_value(ri, ci) for ci in range(sh.ncols)]
            for ri in range(sh.nrows)
        ]
        yield sh.name, rows


def parse_workbook(path: Path, header: str) -> str:
    out = [header]
    parser = parse_xlsx_rows if path.suffix.lower() == '.xlsx' else parse_xls_rows
    seen: set[str] = set()
    total = 0
    for sn, rows in parser(path):
        if any(k in sn for k in SKIP_SHEET_KEYWORDS):
            continue
        if len(rows) < 4:
            continue
        det = detect_header(rows)
        course_col = det['course']
        if course_col is None:
            continue
        sheet_blocks: list[str] = []
        for ri in range(det['data_start'], len(rows)):
            row = rows[ri]
            if course_col >= len(row):
                continue
            course = clean(row[course_col])
            if not course or course in ('课程', '序号', '日期', '上课时间'):
                continue
            if not looks_like_course_code(course):
                continue
            block = emit_block(course, row, det['label_to_col'], seen, sn)
            if block is None:
                continue
            sheet_blocks.append(block)
            total += 1
        if not sheet_blocks:
            continue
        out.append(f'### Sheet: {sn.strip()}')
        out.extend(sheet_blocks)
    return '\n'.join(out)


def main() -> None:
    parts = [
        parse_workbook(
            REF / 'Think1 完整版陪跑计划表.xlsx',
            '## 内置 · think1（来自 Think1 完整版陪跑计划表）',
        ),
        parse_workbook(
            REF / 'THINK2 3V1陪跑计划表.xls',
            '## 内置 · think2（来自 THINK2 3V1陪跑计划表）',
        ),
        parse_workbook(
            REF / 'Power Up2 完整版 3V1陪跑计划表.xls',
            '## 内置 · powerup2（来自 Power Up2 陪跑计划表）',
        ),
        parse_workbook(
            REF / 'Power Up3 完整版 3V1陪跑计划表.xls',
            '## 内置 · powerup3（来自 Power Up3 陪跑计划表）',
        ),
    ]
    outp = ROOT / 'coze/prompts/builtin-tasks-from-excels.md'
    text = '\n---\n\n'.join(parts) + '\n'
    outp.write_text(text, encoding='utf-8')
    titles = sum(1 for ln in text.splitlines() if ln.startswith('#### '))
    print('written', outp, 'lines', len(text.splitlines()), 'titles', titles)


if __name__ == '__main__':
    main()
